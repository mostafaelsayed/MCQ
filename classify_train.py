import cv2
from openpyxl import load_workbook
import os
import numpy as np
import math as m
import matplotlib.pyplot as plt
def find_corners_and_make_image_vertical(img):
	gray = np.float32(img)
	kernel1 = np.ones((51,51),np.uint8)
	gray = cv2.dilate(gray,kernel1,iterations = 1) # dilate to get only a two small points representing the two big circles in the bottom of the image
	corner1,corner2=get_corners(gray) # function to get those circles
	slope = (corner1[1]-corner2[1])/(corner1[0]-corner2[0]) # calculate the slope between these two points
	angle = m.degrees(m.atan(slope)) # calculate the angle
	rows,cols = img.shape
	M = cv2.getRotationMatrix2D((cols/2,rows/2),angle,1) # the matrix to be used to rotate the image
	dst = cv2.warpAffine(img,M,(cols,rows)) # now the image is vertical
	img = dst
	# same steps to get the new position of the two points in the vertical image(except calculating angle of rotation)

	kernel1 = np.ones((51,51),np.uint8)
	gray = cv2.dilate(dst,kernel1,iterations = 1)
	corner1 = get_corners(gray)[0]
	return img,corner1 # return the vertical and the positions of the two big circles
def get_corners(img):
	dst = cv2.cornerHarris(img,15,3,0.04)
	ret, dst = cv2.threshold(dst,0.01*dst.max(),255,0)
	dst = np.uint8(dst)
	ret, labels, stats, centroids = cv2.connectedComponentsWithStats(dst)
	# define the criteria to stop and refine the corners
	criteria = (cv2.TERM_CRITERIA_EPS + cv2.TERM_CRITERIA_MAX_ITER, 100, 0.001)
	corners = cv2.cornerSubPix(img,np.float32(centroids),(5,5),(-1,-1),criteria)
	corners = sorted(corners, key = lambda x: int(x[1]))
	corner1,corner2=corners[-1],corners[-2]
	corners1 = [corner1,corner2]
	corners1 = sorted(corners1, key = lambda x: int(x[0]))
	corner1 = corners1[0]
	corner2 = corners1[1]
	return corner1,corner2
def preprocessing(img): # initial process must be done before the detection process
	vertical_img=find_corners_and_make_image_vertical(img)
	img,p1=vertical_img
	x1=int(p1[0])-15
	x2=int(p1[0])+5*167
	y1=int(p1[1])-780
	y2=int(p1[1])-161
	img=img[y1:y2,x1:x2]
	return img # the image returned shows only the questions
# positions of the four circles in the question
A=[0,40]
B=[40,80]
C=[80,120]
D=[120,160]
answer=[B,C,A,A,D,A,C,C,A,C,A,B,C,C,B,A,D,B,C,B,D,C,D,B,D,C,D,D,B,C,B,B,D,C,B,C,B,C,C,A,B,B,C,C,B] # hold the correct answers positions
def detect(img):
	i1=0
	i2=175
	counter=0
	score=0
	while i1<=830:
		j1=6
		j2=54
	 	while j1<=570:
	 		img1=img[j1:j2,i1:i2]
	 		params=cv2.SimpleBlobDetector_Params()
	 		params.filterByCircularity=1
	 		params.filterByArea=1
	 		params.filterByConvexity=1
	 		params.minArea=100
	 		params.maxArea=1000
	 		params.maxThreshold=1000
	 		params.minThreshold=0
	 		params.minConvexity=0
			params.maxConvexity=1
			params.minCircularity=0
			params.maxCircularity=1
			detector = cv2.SimpleBlobDetector_create(params)
			detected_circles = detector.detect(img1)
			#im_with_detected_circles = cv2.drawdetected_circles(img1, detected_circles, np.array([]), (0,0,255), cv2.DRAW_MATCHES_FLAGS_DRAW_RICH_detected_circles)
			#plt.imshow(im_with_detected_circles,'gray')
			#plt.show()
			if len(detected_circles)==1:
				if detected_circles[0].pt[0]>=answer[counter][0] and detected_circles[0].pt[0]<=answer[counter][1]:
					score=score+1
			elif len(detected_circles)==2:
				if detected_circles[0].pt[0]>=answer[counter][0] and detected_circles[0].pt[0]<=answer[counter][1]:
					average=get_average(img1,detected_circles,1)
				  	if average>190:
				  		score=score+1
				elif detected_circles[1].pt[0]>=answer[counter][0] and detected_circles[1].pt[0]<=answer[counter][1]:
					average=get_average(img1,detected_circles,0)
				  	if average>190:
				  		score=score+1
			elif len(detected_circles)==3:
				if detected_circles[0].pt[0]>=answer[counter][0] and detected_circles[0].pt[0]<=answer[counter][1]:
					average1=get_average(img1,detected_circles,1)
					average2=get_average(img1,detected_circles,2)
				  	if average1>190 and average2>190:
				  		score=score+1
				elif detected_circles[1].pt[0]>=answer[counter][0] and detected_circles[1].pt[0]<=answer[counter][1]:
					average1=get_average(img1,detected_circles,0)
					average2=get_average(img1,detected_circles,2)
				  	if average1>190 and average2>190:
				  		score=score+1
				elif detected_circles[2].pt[0]>=answer[counter][0] and detected_circles[2].pt[0]<=answer[counter][1]:
					average1=get_average(img1,detected_circles,0)
					average2=get_average(img1,detected_circles,1)
				  	if average1>190 and average2>190:
				  		score=score+1
			elif len(detected_circles)==0:
				x=(answer[counter][0]+answer[counter][1])/2
				ran=range(5,24)
				average=[img1[y,x] for i,y in enumerate(ran)]
				average=np.average(average)
				if average<200:
				  		score=score+1
			counter=counter+1
			j1=j1+40
			j2=j2+40
		i1=i1+327
		i2=i2+327
	return score
def get_average(img,detected_circles,j):
	x=int(detected_circles[j].pt[0])
	y=int(detected_circles[j].pt[1])
	ran1=range(x-10,x+10)
	average1=[img[y,x1] for i,x1 in enumerate(ran1)]
	average1=np.average(average1)
	return average1
folder=r'train'
l=[]
def correct():
	row1=2
	for file in os.listdir(folder):
		if file!='image inside train.py':
			cimg=cv2.imread(os.path.join(folder, file),0)
			img=preprocessing(cimg)
			score=detect(img)
			l.append([file,score])
			#print file,score
	wb=load_workbook('mytrain.xlsx')
	ws=wb.active
	while 1:
			cell=ws.cell(row=row1,column=1)
			x=cell.value
			x1=[y for i,y in enumerate(l) if y[0] == x]
			x1=x1[0][1]
			cell1=ws.cell(row=row1,column=2)
			cell1.value=x1
			row1+=1
			if row1==287:
				break
	wb.save('mytrain.xlsx')
correct()